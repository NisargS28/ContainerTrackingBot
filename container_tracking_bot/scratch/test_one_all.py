import sys
import os
import time
from datetime import datetime
import openpyxl

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from carriers.one import ONETracker
from utils.text_utils import extract_date_from_text

def test_one_all():
    excel_path = r"D:\Nisarg.Doc\Skaps\automate\container_tracking\testing_one.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} not found!")
        return

    print(f"Loading {excel_path}...")
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    # Find columns (assuming standard format or just checking headers)
    headers = [cell.value for cell in sheet[1]]
    try:
        cont_col = headers.index("CONTAINER NO.") + 1
    except ValueError:
        try:
            cont_col = headers.index("CONTAINER NO") + 1
        except ValueError:
            cont_col = 2 # Default column 2

    try:
        etd_col = headers.index("ETD") + 1
        eta_col = headers.index("ETA AT PORT") + 1
        status_col = headers.index("STATUS") + 1
        dt_col = headers.index("CURRENT DATE & TIME") + 1
        del_col = headers.index("DELIVERY STATUS") + 1
        move_col = headers.index("LAST MOVE") + 1
    except ValueError:
        # Defaults if headers are slightly different
        etd_col = 3
        eta_col = 4
        status_col = 5
        dt_col = 6
        del_col = 7
        move_col = 8

    tracker = ONETracker()
    print("Tracker Initialized. Browser launched once.\n")

    try:
        for row in range(2, sheet.max_row + 1):
            container_no = sheet.cell(row=row, column=cont_col).value
            if not container_no:
                continue
                
            container_no = str(container_no).strip()
            if not container_no:
                continue

            print(f"Tracking container {container_no} (Row {row})...")
            
            # Add a small delay between requests to be polite
            if row > 2:
                time.sleep(2)
                
            res = tracker.track(container_no)
            
            print(f"Result: {res}\n")
            
            # Write back
            sheet.cell(row=row, column=etd_col, value=res.get("ETD", ""))
            
            eta_val = res.get("ETA", "")
            sheet.cell(row=row, column=eta_col, value=eta_val)
            if eta_val and eta_val != "---":
                parsed = extract_date_from_text(eta_val)
                if parsed:
                    c = sheet.cell(row=row, column=eta_col)
                    c.value = parsed
                    c.number_format = 'dd-mm-yyyy'
                    
            sheet.cell(row=row, column=status_col, value=res.get("TrackingStatus", "Not Tracked"))
            dt_val = res.get("TrackedAt", "")
            dt_cell = sheet.cell(row=row, column=dt_col)
            if dt_val:
                try:
                    parsed_dt = datetime.strptime(dt_val, "%d-%m-%Y %H:%M:%S")
                    dt_cell.value = parsed_dt
                    dt_cell.number_format = 'dd-mm-yyyy hh:mm:ss'
                except Exception:
                    dt_cell.value = dt_val
            else:
                dt_cell.value = ""
            sheet.cell(row=row, column=del_col, value=res.get("DeliveryStatus", ""))
            
            # Write last move if column exists or just put it somewhere
            # If standard format, last move is usually column 3 or 8. 
            sheet.cell(row=row, column=move_col, value=res.get("CurrentStatus", ""))

        print(f"Saving to {excel_path}...")
        try:
            wb.save(excel_path)
            print("Successfully saved to primary file.")
        except PermissionError:
            fallback = excel_path.replace(".xlsx", "_output.xlsx")
            print(f"Primary file locked. Saving to {fallback} instead...")
            wb.save(fallback)
            
    finally:
        tracker.close()
        print("Done!")

if __name__ == "__main__":
    test_one_all()
