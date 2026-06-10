import sys
import os
import time
from datetime import datetime
import openpyxl

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from carriers.hapag import HapagTracker
from utils.text_utils import extract_date_from_text

def run_hapag_excel():
    excel_path = r"D:\Nisarg.Doc\Skaps\automate\container_tracking\testing_hapag.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} not found!")
        return

    print(f"Loading {excel_path}...")
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    # Print sheet headers for debugging
    headers = [cell.value for cell in sheet[1]]
    print(f"Sheet headers: {headers}")

    # Find columns (case-insensitive and matching substrings)
    def find_col(names, default):
        for name in names:
            for i, h in enumerate(headers):
                if h and str(h).strip().upper() == name.upper():
                    return i + 1
        return default

    cont_col = find_col(["CONTAINER NO.", "CONTAINER NO", "CONTAINER"], 2)
    etd_col = find_col(["ETD", "DEPARTURE DATE"], 3)
    eta_col = find_col(["ETA", "ETA AT PORT", "ARRIVAL DATE"], 4)
    status_col = find_col(["STATUS", "TRACKING STATUS"], 5)
    dt_col = find_col(["CURRENT DATE & TIME", "TRACKED AT"], 6)
    del_col = find_col(["DELIVERY STATUS"], 7)
    move_col = find_col(["LAST MOVE", "CURRENT STATUS", "LOCATION"], 8)

    print(f"Using columns -> Container: {cont_col}, ETD: {etd_col}, ETA: {eta_col}, Last Move: {move_col}, Status: {status_col}")

    tracker = HapagTracker()
    print("HapagTracker Initialized. Browser launched once.\n")

    try:
        for row in range(2, sheet.max_row + 1):
            container_no = sheet.cell(row=row, column=cont_col).value
            if not container_no:
                continue
                
            container_no = str(container_no).strip()
            if not container_no:
                continue

            print(f"Tracking container {container_no} (Row {row})...")
            
            # Apply inter-request delay to avoid anti-bot detection
            if row > 2:
                time.sleep(3)
                
            res = tracker.track(container_no)
            print(f"Result: {res}\n")
            
            # Write back tracking info
            # Format dates as date objects if valid
            def write_date_cell(col, date_str):
                cell = sheet.cell(row=row, column=col)
                if date_str and date_str != "---":
                    parsed = extract_date_from_text(date_str)
                    if parsed:
                        cell.value = parsed
                        cell.number_format = 'dd-mm-yyyy'
                    else:
                        cell.value = date_str
                else:
                    cell.value = date_str

            write_date_cell(etd_col, res.get("ETD", ""))
            write_date_cell(eta_col, res.get("ETA", ""))
            
            sheet.cell(row=row, column=status_col, value=res.get("TrackingStatus", "Not Tracked"))
            dt_val = res.get("TrackedAt", "")
            dt_cell = sheet.cell(row=row, column=dt_col)
            if dt_val:
                try:
                    parsed_dt = datetime.strptime(dt_val, "%d-%m-%Y %H:%M:%S")
                    dt_cell.value = parsed_dt
                    dt_cell.number_format = 'dd-mm-yyyy hh:mm:ss'
                except Exception:
                    dt_cell.value = dt_val
            else:
                dt_cell.value = ""
            sheet.cell(row=row, column=del_col, value=res.get("DeliveryStatus", "Not Updated"))
            sheet.cell(row=row, column=move_col, value=res.get("CurrentStatus", ""))

        print(f"Saving to {excel_path}...")
        try:
            wb.save(excel_path)
            print("Successfully saved to primary file.")
        except PermissionError:
            fallback = excel_path.replace(".xlsx", "_output.xlsx")
            print(f"Primary file locked. Saving to {fallback} instead...")
            wb.save(fallback)
            
    finally:
        tracker.close()
        print("Done!")

if __name__ == "__main__":
    run_hapag_excel()
