import os
import shutil
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd
from utils.logger import logger
from utils.text_utils import clean_container_number, normalize_text, extract_date_from_text

def find_column_indices(sheet):
    """
    Scans the first row of a sheet and returns a dictionary mapping 
    logical columns to 1-based column indices.
    """
    row1 = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
    indices = {
        "s_line": None,
        "b_l": None,
        "container_no": None,
        "delivery_status": None,
        "etd": None,
        "eta": None,
        "status": None,
        "tracked_at": None,
        "last_move": None
    }
    
    for idx, cell_val in enumerate(row1, 1):
        if not cell_val:
            continue
        val = str(cell_val).strip().upper()
        
        if ("S/LINE" in val or "S-LINE" in val or "CARRIER" in val) and not any(k in val for k in ("FREIGHT", "CHRG", "MISC", "BILL", "DATE", "DAYS", "HAND", "TIME")):
            indices["s_line"] = idx
        elif "B/L" in val or "BL" in val:
            indices["b_l"] = idx
        elif "CONTAINER NO" in val or "CONTAINER_NO" in val or "CONTAINER" in val:
            indices["container_no"] = idx
        elif "DELIVERY STATUS" in val or "DELIVERY_STATUS" in val:
            indices["delivery_status"] = idx
        elif val == "ETD":
            indices["etd"] = idx
        elif "ETA" in val or "ARRIV" in val:
            indices["eta"] = idx
        elif val == "STATUS":
            indices["status"] = idx
        elif "CURRENT DATE" in val or "TRACKED" in val or "DATE & TIME" in val:
            indices["tracked_at"] = idx
        elif "LAST MOVE" in val or "LAST_MOVE" in val or "MOVE" in val:
            indices["last_move"] = idx

    # If some logical columns were not found, use fallback positions
    if not indices["s_line"]:
        indices["s_line"] = 1
    if not indices["container_no"]:
        # If B/L exists, container_no is usually column 3, otherwise column 2
        indices["container_no"] = 3 if indices["b_l"] else 2
    if not indices["delivery_status"]:
        # If B/L exists, delivery_status is usually column 9, otherwise column 7
        indices["delivery_status"] = 9 if indices["b_l"] else 7
    if not indices["etd"]:
        indices["etd"] = 3
    if not indices["eta"]:
        indices["eta"] = 4
    if not indices["status"]:
        indices["status"] = 5
    if not indices["tracked_at"]:
        indices["tracked_at"] = 6

    return indices

def read_input_workbook(file_path):
    """Loads the workbook using openpyxl."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Input Excel file not found: {file_path}")
    logger.info(f"Opening workbook: {file_path}")
    return openpyxl.load_workbook(file_path, data_only=True)

def extract_msc_rows(wb):
    """
    Loops through all sheets in the workbook, detects columns dynamically,
    and returns a list of MSC tracking candidates.
    Each candidate is represented as a dictionary.
    """
    msc_list = []
    
    for sheet_name in wb.sheetnames:
        # Skip temp sheets or blank sheets
        if sheet_name.lower().startswith("sheet"):
            continue
            
        sheet = wb[sheet_name]
        if sheet.max_row < 2:
            continue
            
        indices = find_column_indices(sheet)
        logger.info(f"Scanning sheet '{sheet_name}'. Columns indices: {indices}")
        
        for row_idx in range(2, sheet.max_row + 1):
            s_line = normalize_text(sheet.cell(row=row_idx, column=indices["s_line"]).value)
            container_no = normalize_text(sheet.cell(row=row_idx, column=indices["container_no"]).value)
            delivery_status = normalize_text(sheet.cell(row=row_idx, column=indices["delivery_status"]).value)
            
            bl_no = ""
            if indices["b_l"] is not None:
                bl_no = normalize_text(sheet.cell(row=row_idx, column=indices["b_l"]).value)
            
            # Checks:
            # 1. Skip if Container No is blank
            if not container_no:
                continue
                
            # 2. Skip row if B/L column exists but B/L is blank
            if indices["b_l"] is not None and not bl_no:
                continue
                
            # 3. S/Line must contain "MSC" case-insensitive
            if "MSC" not in s_line.upper():
                continue
                
            # 4. Delivery Status must be blank or contain "Not" case-insensitive
            del_upper = delivery_status.upper()
            if del_upper and "NOT" not in del_upper:
                continue
                
            # Matches all criteria, add to list
            msc_list.append({
                "SheetName": sheet_name,
                "OriginalExcelRow": row_idx,
                "SLine": s_line,
                "BLNo": bl_no,
                "ContainerNo": clean_container_number(container_no),
                "DeliveryStatus": delivery_status
            })
            
    # Remove duplicates based on ContainerNo + BLNo + SheetName (keep first)
    unique_list = []
    seen = set()
    for row in msc_list:
        key = (row["ContainerNo"], row["BLNo"], row["SheetName"])
        if key not in seen:
            seen.add(key)
            unique_list.append(row)
            
    logger.info(f"Extracted {len(unique_list)} unique MSC containers out of {len(msc_list)} raw matches.")
    return unique_list

def save_msc_list(msc_list, output_path):
    """Saves the filtered MSC list to output/msc_list.xlsx."""
    logger.info(f"Saving filtered MSC list to: {output_path}")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    df = pd.DataFrame(msc_list)
    if df.empty:
        # Create empty DataFrame with expected columns
        df = pd.DataFrame(columns=["SheetName", "OriginalExcelRow", "SLine", "BLNo", "ContainerNo", "DeliveryStatus"])
        
    df.to_excel(output_path, index=False)
    logger.info("MSC list saved successfully.")

def save_tracking_output(tracking_results, output_path):
    """
    Saves the execution results to output/msc_tracking_output.xlsx.
    Includes details and a summary sheet.
    """
    logger.info(f"Saving main tracking output to: {output_path}")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Create details dataframe
    df_details = pd.DataFrame(tracking_results)
    if df_details.empty:
        cols = [
            "SheetName", "OriginalExcelRow", "SLine", "Carrier", "BLNo", "ContainerNo", 
            "DeliveryStatus", "ETD", "ETA", "Vessel", "Voyage", "CurrentStatus", 
            "TrackingStatus", "Remarks", "TrackingURL", "TrackedAt"
        ]
        df_details = pd.DataFrame(columns=cols)
    else:
        # Standardize order of columns
        cols_order = [
            "SheetName", "OriginalExcelRow", "SLine", "Carrier", "BLNo", "ContainerNo", 
            "DeliveryStatus", "ETD", "ETA", "Vessel", "Voyage", "CurrentStatus", 
            "TrackingStatus", "Remarks", "TrackingURL", "TrackedAt"
        ]
        df_details = df_details.reindex(columns=cols_order)
        
    # Calculate Summary Statistics
    total_msc = len(df_details)
    success_count = len(df_details[df_details["TrackingStatus"] == "Success"])
    failed_count = len(df_details[df_details["TrackingStatus"] == "Failed"])
    manual_count = len(df_details[df_details["TrackingStatus"] == "Manual Required"])
    nodata_count = len(df_details[df_details["TrackingStatus"] == "No Data Found"])
    
    summary_data = {
        "Metric": [
            "Total MSC Records",
            "Success",
            "Failed",
            "Manual Required",
            "No Data Found"
        ],
        "Count": [
            total_msc,
            success_count,
            failed_count,
            manual_count,
            nodata_count
        ]
    }
    df_summary = pd.DataFrame(summary_data)
    
    # Write both to sheets
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_details.to_excel(writer, sheet_name="Tracking Details", index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        
    logger.info("Tracking details and summary sheets saved successfully.")

def update_original_copy(tracking_results, input_path, output_path):
    """
    Copies the input file and updates the tracking results.
    Standardizes every sheet's output format to the 7-column layout:
    ['S/LINE', 'CONTAINER NO.', 'ETD', 'ETA AT PORT', 'STATUS', 'CURRENT DATE & TIME', 'DELIVERY STATUS']
    Converts 9-column sheets to this layout by dropping extra columns.
    """
    logger.info(f"Creating copy of original workbook at: {output_path}")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    shutil.copy2(input_path, output_path)
    
    # Open the copy for modification
    wb = openpyxl.load_workbook(output_path)
    
    # Index tracking results for fast lookup by (SheetName, OriginalExcelRow)
    results_map = {(r["SheetName"], r["OriginalExcelRow"]): r for r in tracking_results}
    
    # 7-Column headers and formatting styles
    headers = ['S/LINE', 'CONTAINER NO.', 'ETD', 'ETA AT PORT', 'STATUS', 'CURRENT DATE & TIME', 'DELIVERY STATUS']
    
    for sheet_name in wb.sheetnames:
        if sheet_name.lower().startswith("sheet"):
            continue
            
        sheet = wb[sheet_name]
        logger.info(f"Standardizing and updating sheet: {sheet_name}")
        
        # Read old indices and rows
        indices = find_column_indices(sheet)
        old_max_row = sheet.max_row
        
        # Copy the original rows data in memory first
        original_data_rows = []
        for r in range(2, old_max_row + 1):
            row_data = {
                "row_idx": r,
                "s_line": sheet.cell(row=r, column=indices["s_line"]).value,
                "container_no": sheet.cell(row=r, column=indices["container_no"]).value,
                "etd": sheet.cell(row=r, column=indices["etd"]).value if indices["etd"] <= sheet.max_column else None,
                "eta": sheet.cell(row=r, column=indices["eta"]).value if indices["eta"] <= sheet.max_column else None,
                "status": sheet.cell(row=r, column=indices["status"]).value if indices["status"] <= sheet.max_column else None,
                "tracked_at": sheet.cell(row=r, column=indices["tracked_at"]).value if indices["tracked_at"] <= sheet.max_column else None,
                "delivery_status": sheet.cell(row=r, column=indices["delivery_status"]).value if indices["delivery_status"] <= sheet.max_column else None,
            }
            original_data_rows.append(row_data)
            
        # Clear the old content from the sheet completely
        sheet.delete_rows(1, sheet.max_row + 1)
        
        # Write new headers
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            # Apply standard bold style for headers
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Write rows back, merging with tracking results
        for i, old_row in enumerate(original_data_rows, 2):
            r_idx = old_row["row_idx"]
            
            # Check if we have tracking results for this sheet and row
            res = results_map.get((sheet_name, r_idx))
            
            if res:
                # S/LINE
                s_line_val = res.get("SLine") or old_row["s_line"]
                # CONTAINER NO
                container_val = res.get("ContainerNo") or old_row["container_no"]
                # ETD
                etd_val = res.get("ETD") or ""
                # ETA
                eta_val = res.get("ETA") or ""
                # STATUS
                status_val = res.get("TrackingStatus")
                if status_val == "Success":
                    status_val = "Tracked"
                elif status_val == "Manual Required":
                    status_val = "Manual Required"
                elif status_val == "No Data Found":
                    status_val = "No Data Found"
                else:
                    status_val = "Not Tracked"
                
                # CURRENT DATE & TIME
                tracked_at_val = res.get("TrackedAt")
                
                # DELIVERY STATUS
                delivery_val = res.get("DeliveryStatus")
            else:
                # Keep original data
                s_line_val = old_row["s_line"]
                container_val = old_row["container_no"]
                etd_val = old_row["etd"]
                eta_val = old_row["eta"]
                status_val = old_row["status"]
                tracked_at_val = old_row["tracked_at"]
                delivery_val = old_row["delivery_status"]
                
            # Write to cells
            sheet.cell(row=i, column=1, value=s_line_val)
            sheet.cell(row=i, column=2, value=container_val)
            sheet.cell(row=i, column=3, value=etd_val)
            sheet.cell(row=i, column=4, value=eta_val)
            sheet.cell(row=i, column=5, value=status_val)
            sheet.cell(row=i, column=6, value=tracked_at_val)
            sheet.cell(row=i, column=7, value=delivery_val)
            
            # Format dates nicely
            for col in (3, 4):
                cell = sheet.cell(row=i, column=col)
                if isinstance(cell.value, datetime):
                    cell.number_format = 'dd-mm-yyyy'
                elif isinstance(cell.value, str) and cell.value:
                    # Try to parse string to format
                    parsed = extract_date_from_text(cell.value)
                    if parsed:
                        cell.value = parsed
                        cell.number_format = 'dd-mm-yyyy'
            
            # Format Tracked At datetime
            cell_tracked = sheet.cell(row=i, column=6)
            if isinstance(cell_tracked.value, datetime):
                cell_tracked.number_format = 'dd-mm-yyyy hh:mm:ss'
            elif isinstance(cell_tracked.value, str) and cell_tracked.value:
                try:
                    parsed = datetime.strptime(cell_tracked.value, '%d-%m-%Y %H:%M:%S')
                    cell_tracked.value = parsed
                    cell_tracked.number_format = 'dd-mm-yyyy hh:mm:ss'
                except ValueError:
                    try:
                        parsed = datetime.strptime(cell_tracked.value, '%Y-%m-%d %H:%M:%S')
                        cell_tracked.value = parsed
                        cell_tracked.number_format = 'dd-mm-yyyy hh:mm:ss'
                    except ValueError:
                        pass

        # Autofit columns
        for col in sheet.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(output_path)
    logger.info("Updated original copy saved and standardized successfully.")


# ============================================================
# NEW PER-SHEET FUNCTIONS (used by the updated main.py flow)
# ============================================================

def sync_input_to_output(input_path, output_path, sheet_name):
    """
    Syncs S/LINE and CONTAINER NO from the input sheet to the output sheet in standard 8-column format:
    S/LINE | CONTAINER NO. | LAST MOVE | ETD | ETA AT PORT | STATUS | CURRENT DATE & TIME | DELIVERY STATUS
    Returns True on success, False if file is locked or saving fails.
    """
    if not os.path.exists(input_path):
        logger.error(f"Input file not found for sync: {input_path}")
        return False

    # 1. Load input sheet and read S/LINE & CONTAINER NO
    try:
        in_wb = openpyxl.load_workbook(input_path, data_only=True)
        if sheet_name not in in_wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in input workbook. Skipping sync.")
            return True
        in_ws = in_wb[sheet_name]
    except Exception as e:
        logger.error(f"Error loading input workbook for sync: {e}")
        return False

    in_indices = find_column_indices(in_ws)
    in_sline_col = in_indices["s_line"]
    in_cont_col = in_indices["container_no"]

    if not in_sline_col or not in_cont_col:
        logger.warning(f"Could not find S/LINE or CONTAINER NO columns in input sheet '{sheet_name}'. Skipping sync.")
        return True

    # Read rows from input
    input_entries = []
    seen_in_clean = set()
    for r in range(2, in_ws.max_row + 1):
        s_line = in_ws.cell(row=r, column=in_sline_col).value
        cont_no = in_ws.cell(row=r, column=in_cont_col).value
        
        if not cont_no:
            continue
            
        cleaned_no = clean_container_number(cont_no)
        if not cleaned_no:
            continue
            
        s_line_str = str(s_line).strip() if s_line else ""
        
        if cleaned_no not in seen_in_clean:
            seen_in_clean.add(cleaned_no)
            input_entries.append((s_line_str, str(cont_no).strip(), cleaned_no))

    # 2. Load or create output workbook
    if os.path.exists(output_path):
        try:
            out_wb = openpyxl.load_workbook(output_path)
        except Exception as e:
            logger.error(f"Error loading output workbook for sync: {e}")
            return False
    else:
        out_wb = openpyxl.Workbook()
        if "Sheet" in out_wb.sheetnames:
            del out_wb["Sheet"]

    if sheet_name in out_wb.sheetnames:
        out_ws = out_wb[sheet_name]
    else:
        out_ws = out_wb.create_sheet(title=sheet_name)

    # 3. Read existing rows in output sheet (if any)
    existing_rows_data = []
    headers = [
        "S/LINE", "CONTAINER NO.", "LAST MOVE", "ETD", "ETA AT PORT",
        "STATUS", "CURRENT DATE & TIME", "DELIVERY STATUS"
    ]
    
    if out_ws.max_row >= 2:
        out_indices = find_column_indices(out_ws)
        for r in range(2, out_ws.max_row + 1):
            row_data = {
                "s_line": out_ws.cell(row=r, column=out_indices["s_line"]).value if out_indices["s_line"] and out_indices["s_line"] <= out_ws.max_column else None,
                "container_no": out_ws.cell(row=r, column=out_indices["container_no"]).value if out_indices["container_no"] and out_indices["container_no"] <= out_ws.max_column else None,
                "last_move": out_ws.cell(row=r, column=out_indices["last_move"]).value if out_indices.get("last_move") and out_indices["last_move"] <= out_ws.max_column else None,
                "etd": out_ws.cell(row=r, column=out_indices["etd"]).value if out_indices["etd"] and out_indices["etd"] <= out_ws.max_column else None,
                "eta": out_ws.cell(row=r, column=out_indices["eta"]).value if out_indices["eta"] and out_indices["eta"] <= out_ws.max_column else None,
                "status": out_ws.cell(row=r, column=out_indices["status"]).value if out_indices["status"] and out_indices["status"] <= out_ws.max_column else None,
                "tracked_at": out_ws.cell(row=r, column=out_indices["tracked_at"]).value if out_indices["tracked_at"] and out_indices["tracked_at"] <= out_ws.max_column else None,
                "delivery_status": out_ws.cell(row=r, column=out_indices["delivery_status"]).value if out_indices["delivery_status"] and out_indices["delivery_status"] <= out_ws.max_column else None,
            }
            existing_rows_data.append(row_data)

    # 4. Clear output sheet and write headers
    out_ws.delete_rows(1, out_ws.max_row + 1)
    
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = out_ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Map of output containers
    out_map = {clean_container_number(r["container_no"]): r for r in existing_rows_data if r["container_no"]}
    
    synced_rows = []
    updated_containers = set()
    
    # First, process input entries
    for s_line_val, raw_cont_val, clean_cont_val in input_entries:
        if clean_cont_val in out_map:
            # Update S/Line and Container No, keep others
            orig_row = out_map[clean_cont_val]
            orig_row["s_line"] = s_line_val
            orig_row["container_no"] = raw_cont_val
            synced_rows.append(orig_row)
            updated_containers.add(clean_cont_val)
        else:
            # Append new row
            new_row = {
                "s_line": s_line_val,
                "container_no": raw_cont_val,
                "last_move": "",
                "etd": "",
                "eta": "",
                "status": "Not Tracked",
                "tracked_at": "",
                "delivery_status": ""
            }
            synced_rows.append(new_row)
            updated_containers.add(clean_cont_val)
            
    # Write back any containers that were in output but NOT in input
    for clean_cont_val, orig_row in out_map.items():
        if clean_cont_val not in updated_containers:
            synced_rows.append(orig_row)

    # Write all rows to the sheet
    for row_idx, r_data in enumerate(synced_rows, 2):
        out_ws.cell(row=row_idx, column=1, value=r_data["s_line"])
        out_ws.cell(row=row_idx, column=2, value=r_data["container_no"])
        out_ws.cell(row=row_idx, column=3, value=r_data["last_move"])
        out_ws.cell(row=row_idx, column=4, value=r_data["etd"])
        out_ws.cell(row=row_idx, column=5, value=r_data["eta"])
        out_ws.cell(row=row_idx, column=6, value=r_data["status"])
        out_ws.cell(row=row_idx, column=7, value=r_data["tracked_at"])
        out_ws.cell(row=row_idx, column=8, value=r_data["delivery_status"])
        
        # Format ETD/ETA
        for col in (4, 5):
            cell = out_ws.cell(row=row_idx, column=col)
            if isinstance(cell.value, datetime):
                cell.number_format = "dd-mm-yyyy"
            elif isinstance(cell.value, str) and cell.value:
                parsed = extract_date_from_text(cell.value)
                if parsed:
                    cell.value = parsed
                    cell.number_format = "dd-mm-yyyy"
                    
        cell_dt = out_ws.cell(row=row_idx, column=7)
        if isinstance(cell_dt.value, str) and cell_dt.value:
            try:
                parsed_dt = datetime.strptime(cell_dt.value, "%d-%m-%Y %H:%M:%S")
                cell_dt.value = parsed_dt
            except ValueError:
                try:
                    parsed_dt = datetime.strptime(cell_dt.value, "%Y-%m-%d %H:%M:%S")
                    cell_dt.value = parsed_dt
                except ValueError:
                    pass
        if isinstance(cell_dt.value, datetime):
            cell_dt.number_format = "dd-mm-yyyy hh:mm:ss"

    # Autofit columns
    for col in out_ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        out_ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    try:
        out_wb.save(output_path)
        logger.info(f"Sync complete and saved for sheet '{sheet_name}'.")
        return True
    except PermissionError:
        logger.error(
            f"PermissionError: '{output_path}' is open in Excel or locked. "
            f"Please close the file and re-run. Sync for '{sheet_name}' was NOT saved!"
        )
        return False


def extract_candidates_for_carrier(output_path, sheet_name, carrier_keyword):
    """
    Opens the output file for sheet_name and extracts container numbers
    where S/LINE contains carrier_keyword (case-insensitive) and DELIVERY STATUS
    is not 'Delivered'.
    """
    if not os.path.exists(output_path):
        return []

    try:
        wb = openpyxl.load_workbook(output_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
    except Exception as e:
        logger.error(f"Error loading output file for extracting candidates: {e}")
        return []

    indices = find_column_indices(ws)
    s_line_col = indices["s_line"]
    container_col = indices["container_no"]
    delivery_col = indices["delivery_status"]

    if not s_line_col or not container_col or not delivery_col:
        logger.warning(f"Required columns missing in output sheet '{sheet_name}'. Cannot extract candidates.")
        return []

    candidates = []
    for r in range(2, ws.max_row + 1):
        s_line_val = ws.cell(row=r, column=s_line_col).value
        container_val = ws.cell(row=r, column=container_col).value
        delivery_val = ws.cell(row=r, column=delivery_col).value

        s_line_str = str(s_line_val or "").strip()
        container_str = str(container_val or "").strip()
        delivery_str = str(delivery_val or "").strip()

        if carrier_keyword.upper() not in s_line_str.upper():
            continue

        cleaned_no = clean_container_number(container_str)
        if not cleaned_no:
            continue

        del_upper = delivery_str.upper().replace(" ", "")
        if del_upper == "DELIVERED":
            continue

        candidates.append({
            "SheetName": sheet_name,
            "OriginalExcelRow": r,
            "SLine": s_line_str,
            "ContainerNo": cleaned_no,
            "BLNo": "",
            "DeliveryStatus": delivery_str
        })

    logger.info(f"Sheet '{sheet_name}': Found {len(candidates)} {carrier_keyword} candidates to track (Delivery Status is not Delivered).")
    return candidates


def write_results_to_reference(tracking_results, reference_path, sheet_name):
    """
    Writes tracking results for one sheet into Copy of bl_full_with delivery.xlsx
    matching by Container Number.
    """
    if not tracking_results:
        logger.warning(f"No results for '{sheet_name}' — nothing written to reference file.")
        return

    logger.info(
        f"Writing {len(tracking_results)} results for sheet '{sheet_name}' "
        f"to reference file: {reference_path}"
    )

    if not os.path.exists(reference_path):
        logger.error(f"Reference file '{reference_path}' does not exist for write-back. Skipping.")
        return

    try:
        wb = openpyxl.load_workbook(reference_path)
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in reference workbook. Skipping.")
            return
        ws = wb[sheet_name]
    except Exception as e:
        logger.error(f"Error loading reference workbook: {e}")
        return

    # Read existing rows in output sheet
    existing_rows_data = []
    indices = find_column_indices(ws)
    for r in range(2, ws.max_row + 1):
        row_data = {
            "s_line": ws.cell(row=r, column=indices["s_line"]).value if indices["s_line"] and indices["s_line"] <= ws.max_column else None,
            "container_no": ws.cell(row=r, column=indices["container_no"]).value if indices["container_no"] and indices["container_no"] <= ws.max_column else None,
            "last_move": ws.cell(row=r, column=indices["last_move"]).value if indices.get("last_move") and indices["last_move"] <= ws.max_column else None,
            "etd": ws.cell(row=r, column=indices["etd"]).value if indices["etd"] and indices["etd"] <= ws.max_column else None,
            "eta": ws.cell(row=r, column=indices["eta"]).value if indices["eta"] and indices["eta"] <= ws.max_column else None,
            "status": ws.cell(row=r, column=indices["status"]).value if indices["status"] and indices["status"] <= ws.max_column else None,
            "tracked_at": ws.cell(row=r, column=indices["tracked_at"]).value if indices["tracked_at"] and indices["tracked_at"] <= ws.max_column else None,
            "delivery_status": ws.cell(row=r, column=indices["delivery_status"]).value if indices["delivery_status"] and indices["delivery_status"] <= ws.max_column else None,
        }
        existing_rows_data.append(row_data)

    # Clear target sheet and write headers
    ws.delete_rows(1, ws.max_row + 1)

    headers = [
        "S/LINE", "CONTAINER NO.", "LAST MOVE", "ETD", "ETA AT PORT",
        "STATUS", "CURRENT DATE & TIME", "DELIVERY STATUS",
    ]

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Mapping of container number to its tracked result
    tracked_map = {clean_container_number(res["ContainerNo"]): res for res in tracking_results if res.get("ContainerNo")}
    
    status_map = {
        "Success":         "Tracked",
        "Failed":          "Not Tracked",
        "Manual Required": "Manual Required",
        "No Data Found":   "No Data Found",
        "Skipped (Blank)": "Skipped",
    }

    def write_row_data(row_idx, s_line_val, container_val, last_move_val, etd_val, eta_val, status_val, tracked_at_val, delivery_val):
        ws.cell(row=row_idx, column=1, value=s_line_val)
        ws.cell(row=row_idx, column=2, value=container_val)
        ws.cell(row=row_idx, column=3, value=last_move_val)
        ws.cell(row=row_idx, column=4, value=etd_val)
        ws.cell(row=row_idx, column=5, value=eta_val)
        ws.cell(row=row_idx, column=6, value=status_val)
        ws.cell(row=row_idx, column=7, value=tracked_at_val)
        ws.cell(row=row_idx, column=8, value=delivery_val)

        # Format ETD/ETA
        for col in (4, 5):
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell.value, datetime):
                cell.number_format = "dd-mm-yyyy"
            elif isinstance(cell.value, str) and cell.value:
                parsed = extract_date_from_text(cell.value)
                if parsed:
                    cell.value = parsed
                    cell.number_format = "dd-mm-yyyy"

        # Format TrackedAt datetime
        cell_dt = ws.cell(row=row_idx, column=7)
        if isinstance(cell_dt.value, str) and cell_dt.value:
            try:
                parsed_dt = datetime.strptime(cell_dt.value, "%d-%m-%Y %H:%M:%S")
                cell_dt.value = parsed_dt
            except ValueError:
                try:
                    parsed_dt = datetime.strptime(cell_dt.value, "%Y-%m-%d %H:%M:%S")
                    cell_dt.value = parsed_dt
                except ValueError:
                    pass
        if isinstance(cell_dt.value, datetime):
            cell_dt.number_format = "dd-mm-yyyy hh:mm:ss"

    write_row = 2
    for old_row in existing_rows_data:
        container_no = old_row["container_no"]
        cleaned_no = clean_container_number(container_no) if container_no else ""

        if cleaned_no and cleaned_no in tracked_map:
            res = tracked_map[cleaned_no]
            status_display = status_map.get(res.get("TrackingStatus", ""), "Not Tracked")
            
            write_row_data(
                row_idx=write_row,
                s_line_val=res.get("SLine", "") or old_row["s_line"],
                container_val=res.get("ContainerNo", "") or old_row["container_no"],
                last_move_val=res.get("CurrentStatus", "") or old_row["last_move"] or "",
                etd_val=res.get("ETD", "") or old_row["etd"],
                eta_val=res.get("ETA", "") or old_row["eta"],
                status_val=status_display,
                tracked_at_val=res.get("TrackedAt", "") or old_row["tracked_at"],
                delivery_val=res.get("DeliveryStatus", "") or old_row["delivery_status"]
            )
        else:
            write_row_data(
                row_idx=write_row,
                s_line_val=old_row["s_line"],
                container_val=old_row["container_no"],
                last_move_val=old_row["last_move"] or "",
                etd_val=old_row["etd"],
                eta_val=old_row["eta"],
                status_val=old_row["status"],
                tracked_at_val=old_row["tracked_at"],
                delivery_val=old_row["delivery_status"]
            )
        write_row += 1

    # Autofit columns
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    try:
        wb.save(reference_path)
        logger.info(f"Reference file updated successfully for sheet '{sheet_name}': {reference_path}")
    except PermissionError:
        logger.error(
            f"PermissionError: '{reference_path}' is open in Excel or locked. "
            f"Please close the file and re-run. Results for '{sheet_name}' were NOT saved!"
        )

